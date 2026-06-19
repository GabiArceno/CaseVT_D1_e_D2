import csv
import os

COLUNAS_ENTRADA = [
    'nome', 'sobrenome', 'data_nascimento', 'telefone', 'email', 'cep',
    'codigo_produto', 'nome_produto', 'descricao_produto', 'valor_produto', 'quantidade',
]
COLUNAS_SAIDA = ['numero_pedido', 'status']


def _detectar_delimitador(caminho):
    with open(caminho, newline='', encoding='utf-8') as f:
        amostra = f.read(2048)
    try:
        return csv.Sniffer().sniff(amostra, delimiters=',;\t').delimiter
    except csv.Error:
        return ','


def ler_pedidos(caminho):
    ext = os.path.splitext(caminho)[1].lower()
    if ext == '.xlsx':
        return _ler_xlsx(caminho)
    return _ler_texto(caminho, ext)


def _ler_texto(caminho, ext):
    delimitador = _detectar_delimitador(caminho) if ext == '.txt' else ','
    with open(caminho, newline='', encoding='utf-8') as f:
        leitor = csv.DictReader(f, delimiter=delimitador)
        linhas = [dict(linha) for linha in leitor]
    for linha in linhas:
        for coluna in COLUNAS_SAIDA:
            linha.setdefault(coluna, '')
    return linhas


def _ler_xlsx(caminho):
    from openpyxl import load_workbook
    wb = load_workbook(caminho)
    ws = wb.active
    linhas_brutas = list(ws.iter_rows(values_only=True))
    cabecalho = [str(c).strip() if c is not None else '' for c in linhas_brutas[0]]
    linhas = []
    for valores in linhas_brutas[1:]:
        if all(v is None for v in valores):
            continue
        linha = {cabecalho[i]: ('' if valores[i] is None else valores[i]) for i in range(len(cabecalho))}
        for coluna in COLUNAS_SAIDA:
            linha.setdefault(coluna, '')
        linhas.append(linha)
    return linhas


def atualizar_pedido(caminho, indice, numero_pedido, status):
    ext = os.path.splitext(caminho)[1].lower()
    if ext == '.xlsx':
        _atualizar_xlsx(caminho, indice, numero_pedido, status)
    else:
        _atualizar_texto(caminho, ext, indice, numero_pedido, status)


def _atualizar_texto(caminho, ext, indice, numero_pedido, status):
    linhas = _ler_texto(caminho, ext)
    linhas[indice]['numero_pedido'] = numero_pedido
    linhas[indice]['status'] = status
    delimitador = _detectar_delimitador(caminho) if ext == '.txt' else ','
    colunas = list(linhas[0].keys()) if linhas else COLUNAS_ENTRADA + COLUNAS_SAIDA
    with open(caminho, 'w', newline='', encoding='utf-8') as f:
        escritor = csv.DictWriter(f, fieldnames=colunas, delimiter=delimitador)
        escritor.writeheader()
        escritor.writerows(linhas)


def _atualizar_xlsx(caminho, indice, numero_pedido, status):
    from openpyxl import load_workbook
    wb = load_workbook(caminho)
    ws = wb.active
    cabecalho = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if 'numero_pedido' not in cabecalho:
        col_pedido = len(cabecalho) + 1
        col_status = col_pedido + 1
        ws.cell(row=1, column=col_pedido, value='numero_pedido')
        ws.cell(row=1, column=col_status, value='status')
    else:
        col_pedido = cabecalho.index('numero_pedido') + 1
        col_status = cabecalho.index('status') + 1
    linha_planilha = indice + 2  # +1 cabeçalho, +1 índice 1-based do openpyxl
    ws.cell(row=linha_planilha, column=col_pedido, value=numero_pedido)
    ws.cell(row=linha_planilha, column=col_status, value=status)
    wb.save(caminho)
